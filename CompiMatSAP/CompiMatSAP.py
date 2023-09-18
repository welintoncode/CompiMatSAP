from PyQt6 import QtCore, QtGui, QtWidgets

import pandas as pd

import sys
from PyQt6.QtWidgets import (
    QTableWidgetItem, QHeaderView
)

from os import system
import openpyxl

import pyperclip as cp

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.setEnabled(True)
        MainWindow.resize(784, 637)
        MainWindow.setStyleSheet("#statusbar{\n"
"    background-color: rgb(43, 43, 43);\n"
"    color: rgb(255, 255, 255);\n"
"    font: 700 12pt \"Segoe UI\";\n"
"}\n"
"\n"
"")
        self.centralwidget = QtWidgets.QWidget(parent=MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.tblMateriales = QtWidgets.QTableWidget(parent=self.centralwidget)
        self.tblMateriales.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        self.tblMateriales.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.tblMateriales.setObjectName("tblMateriales")
        self.tblMateriales.setColumnCount(3)
        self.tblMateriales.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblMateriales.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblMateriales.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblMateriales.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tblMateriales.setHorizontalHeaderItem(2, item)
        self.tblMateriales.horizontalHeader().setStretchLastSection(True)
        self.tblMateriales.verticalHeader().setCascadingSectionResizes(False)
        self.tblMateriales.verticalHeader().setSortIndicatorShown(False)
        self.tblMateriales.verticalHeader().setStretchLastSection(False)
        self.horizontalLayout.addWidget(self.tblMateriales)
        self.tblLetras = QtWidgets.QTableWidget(parent=self.centralwidget)
        self.tblLetras.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.tblLetras.setObjectName("tblLetras")
        self.tblLetras.setColumnCount(3)
        self.tblLetras.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblLetras.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        item.setBackground(QtGui.QColor(255, 255, 255))
        self.tblLetras.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblLetras.setHorizontalHeaderItem(2, item)
        self.tblLetras.horizontalHeader().setStretchLastSection(True)
        self.horizontalLayout.addWidget(self.tblLetras)
        self.gridLayout.addLayout(self.horizontalLayout, 3, 0, 1, 2)
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setSpacing(1)
        self.verticalLayout.setObjectName("verticalLayout")
        self.label_5 = QtWidgets.QLabel(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        self.label_5.setFont(font)
        self.label_5.setStyleSheet("color: rgb(255, 255, 255);\n"
"background-color: rgb(90, 156, 176);")
        self.label_5.setAlignment(QtCore.Qt.AlignmentFlag.AlignBottom|QtCore.Qt.AlignmentFlag.AlignHCenter)
        self.label_5.setWordWrap(False)
        self.label_5.setIndent(-1)
        self.label_5.setOpenExternalLinks(False)
        self.label_5.setObjectName("label_5")
        self.verticalLayout.addWidget(self.label_5)
        self.tblCompilado = QtWidgets.QTableWidget(parent=self.centralwidget)
        self.tblCompilado.setAutoFillBackground(False)
        self.tblCompilado.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        self.tblCompilado.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectItems)
        self.tblCompilado.setObjectName("tblCompilado")
        self.tblCompilado.setColumnCount(12)
        self.tblCompilado.setRowCount(0)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        font.setItalic(False)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(9, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(10, item)
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeading|QtCore.Qt.AlignmentFlag.AlignVCenter)
        font = QtGui.QFont()
        font.setBold(True)
        item.setFont(font)
        self.tblCompilado.setHorizontalHeaderItem(11, item)
        self.tblCompilado.horizontalHeader().setCascadingSectionResizes(False)
        self.tblCompilado.horizontalHeader().setHighlightSections(True)
        self.tblCompilado.horizontalHeader().setSortIndicatorShown(False)
        self.tblCompilado.horizontalHeader().setStretchLastSection(True)
        self.verticalLayout.addWidget(self.tblCompilado)
        self.gridLayout.addLayout(self.verticalLayout, 6, 0, 1, 2)
        self.horizontalLayout_5 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        self.txBuscarLetras = QtWidgets.QLineEdit(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(11)
        self.txBuscarLetras.setFont(font)
        self.txBuscarLetras.setObjectName("txBuscarLetras")
        self.horizontalLayout_5.addWidget(self.txBuscarLetras)
        self.btnLimpiarLetras = QtWidgets.QPushButton(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        self.btnLimpiarLetras.setFont(font)
        self.btnLimpiarLetras.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.btnLimpiarLetras.setObjectName("btnLimpiarLetras")
        self.horizontalLayout_5.addWidget(self.btnLimpiarLetras)
        self.btnBuscarLetras = QtWidgets.QPushButton(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        self.btnBuscarLetras.setFont(font)
        self.btnBuscarLetras.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.btnBuscarLetras.setObjectName("btnBuscarLetras")
        self.horizontalLayout_5.addWidget(self.btnBuscarLetras)
        self.gridLayout.addLayout(self.horizontalLayout_5, 1, 1, 1, 1)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Minimum)
        self.horizontalLayout_2.addItem(spacerItem)
        self.btnPasarMateriales = QtWidgets.QPushButton(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setBold(True)
        font.setUnderline(False)
        font.setStrikeOut(False)
        self.btnPasarMateriales.setFont(font)
        self.btnPasarMateriales.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.btnPasarMateriales.setStyleSheet("")
        self.btnPasarMateriales.setText("")
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("flecha.png"), QtGui.QIcon.Mode.Normal, QtGui.QIcon.State.Off)
        self.btnPasarMateriales.setIcon(icon)
        self.btnPasarMateriales.setObjectName("btnPasarMateriales")
        self.horizontalLayout_2.addWidget(self.btnPasarMateriales)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Minimum)
        self.horizontalLayout_2.addItem(spacerItem1)
        self.btnPasarLetras = QtWidgets.QPushButton(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setBold(True)
        font.setUnderline(False)
        font.setStrikeOut(False)
        self.btnPasarLetras.setFont(font)
        self.btnPasarLetras.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.btnPasarLetras.setStyleSheet("")
        self.btnPasarLetras.setText("")
        self.btnPasarLetras.setIcon(icon)
        self.btnPasarLetras.setObjectName("btnPasarLetras")
        self.horizontalLayout_2.addWidget(self.btnPasarLetras)
        self.gridLayout.addLayout(self.horizontalLayout_2, 0, 0, 1, 2)
        self.horizontalLayout_6 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_6.setObjectName("horizontalLayout_6")
        self.label = QtWidgets.QLabel(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        self.label.setFont(font)
        self.label.setStyleSheet("background-color: rgb(67, 67, 67);\n"
"color: rgb(255, 255, 255);")
        self.label.setAlignment(QtCore.Qt.AlignmentFlag.AlignBottom|QtCore.Qt.AlignmentFlag.AlignHCenter)
        self.label.setWordWrap(False)
        self.label.setIndent(-1)
        self.label.setOpenExternalLinks(False)
        self.label.setObjectName("label")
        self.horizontalLayout_6.addWidget(self.label)
        self.label_2 = QtWidgets.QLabel(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        self.label_2.setFont(font)
        self.label_2.setStyleSheet("background-color: rgb(67, 67, 67);\n"
"color: rgb(255, 255, 255);")
        self.label_2.setAlignment(QtCore.Qt.AlignmentFlag.AlignBottom|QtCore.Qt.AlignmentFlag.AlignHCenter)
        self.label_2.setWordWrap(False)
        self.label_2.setIndent(-1)
        self.label_2.setOpenExternalLinks(False)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout_6.addWidget(self.label_2)
        self.gridLayout.addLayout(self.horizontalLayout_6, 2, 0, 1, 2)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.btnExcel = QtWidgets.QPushButton(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        self.btnExcel.setFont(font)
        self.btnExcel.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.btnExcel.setObjectName("btnExcel")
        self.horizontalLayout_3.addWidget(self.btnExcel)
        self.btnEliminar = QtWidgets.QPushButton(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        self.btnEliminar.setFont(font)
        self.btnEliminar.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.btnEliminar.setObjectName("btnEliminar")
        self.horizontalLayout_3.addWidget(self.btnEliminar)
        self.btnCopiar = QtWidgets.QPushButton(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        self.btnCopiar.setFont(font)
        self.btnCopiar.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.btnCopiar.setObjectName("btnCopiar")
        self.horizontalLayout_3.addWidget(self.btnCopiar)
        self.btnLimpiar = QtWidgets.QPushButton(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        self.btnLimpiar.setFont(font)
        self.btnLimpiar.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.btnLimpiar.setObjectName("btnLimpiar")
        self.horizontalLayout_3.addWidget(self.btnLimpiar)
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Minimum)
        self.horizontalLayout_3.addItem(spacerItem2)
        self.cbAlamcen = QtWidgets.QComboBox(parent=self.centralwidget)
        self.cbAlamcen.setEnabled(True)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.cbAlamcen.setFont(font)
        self.cbAlamcen.setObjectName("cbAlamcen")
        self.cbAlamcen.addItem("")
        self.cbAlamcen.addItem("")
        self.horizontalLayout_3.addWidget(self.cbAlamcen)
        self.gridLayout.addLayout(self.horizontalLayout_3, 5, 0, 1, 2)
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        self.label_3 = QtWidgets.QLabel(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.horizontalLayout_4.addWidget(self.label_3)
        self.txBuscarMateriales = QtWidgets.QLineEdit(parent=self.centralwidget)
        font = QtGui.QFont()
        font.setPointSize(11)
        self.txBuscarMateriales.setFont(font)
        self.txBuscarMateriales.setObjectName("txBuscarMateriales")
        self.horizontalLayout_4.addWidget(self.txBuscarMateriales)
        self.gridLayout.addLayout(self.horizontalLayout_4, 1, 0, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(parent=MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 784, 22))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(parent=MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "COMPILADOR DE MATERIALES"))
        item = self.tblMateriales.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "CODIGO"))
        item = self.tblMateriales.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "DESCRIPCION"))
        item = self.tblMateriales.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "CANTIDAD"))
        item = self.tblLetras.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "CODIGO"))
        item = self.tblLetras.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "LETRA"))
        item = self.tblLetras.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "CANTIDAD"))
        self.label_5.setText(_translate("MainWindow", "COMPILADO DE MATERIALES"))
        item = self.tblCompilado.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Codigo"))
        item = self.tblCompilado.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Descripcion"))
        item = self.tblCompilado.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Cantidad"))
        item = self.tblCompilado.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "UM"))
        item = self.tblCompilado.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "TP"))
        item = self.tblCompilado.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "Stock especial"))
        item = self.tblCompilado.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "Alm."))
        item = self.tblCompilado.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "Ce."))
        item = self.tblCompilado.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "Op."))
        item = self.tblCompilado.horizontalHeaderItem(9)
        item.setText(_translate("MainWindow", "Lote"))
        item = self.tblCompilado.horizontalHeaderItem(10)
        item.setText(_translate("MainWindow", "Tipo aprovis."))
        item = self.tblCompilado.horizontalHeaderItem(11)
        item.setText(_translate("MainWindow", "Destinatario"))
        self.btnLimpiarLetras.setText(_translate("MainWindow", "C"))
        self.btnBuscarLetras.setText(_translate("MainWindow", "Compilar"))
        self.label.setText(_translate("MainWindow", "TABLA MATERIALES"))
        self.label_2.setText(_translate("MainWindow", "TABLA LETRAS"))
        self.btnExcel.setText(_translate("MainWindow", "EXCEL"))
        self.btnEliminar.setText(_translate("MainWindow", "Eliminar"))
        self.btnCopiar.setText(_translate("MainWindow", "Copiar"))
        self.btnLimpiar.setText(_translate("MainWindow", "Limpiar"))
        self.cbAlamcen.setItemText(0, _translate("MainWindow", "D015"))
        self.cbAlamcen.setItemText(1, _translate("MainWindow", "D016"))
        self.label_3.setText(_translate("MainWindow", "Buscar"))



#FUNCIONES AGREGADAS

        self.almacen = 'D015'
        #self.almacen = self.cbAlamcen.currentText()

        #MODIFICACION COLUMNAS
        self.tblMateriales.setColumnWidth(0, 60)
        self.tblMateriales.setColumnWidth(1, 200)
        self.tblMateriales.setColumnWidth(2, 70)

        self.tblLetras.setColumnWidth(0, 96)
        self.tblLetras.setColumnWidth(1, 125)
        self.tblLetras.setColumnWidth(2, 96)
        
        
        self.tblCompilado.setColumnWidth(0, 60)
        self.tblCompilado.setColumnWidth(1, 370)
        self.tblCompilado.setColumnWidth(2, 70)
        self.tblCompilado.setColumnWidth(3, 25)
        self.tblCompilado.setColumnWidth(4, 25)
        self.tblCompilado.setColumnWidth(5, 25)
        self.tblCompilado.setColumnWidth(6, 40)
        self.tblCompilado.setColumnWidth(7, 50)
        self.tblCompilado.setColumnWidth(8, 50)
        self.tblCompilado.setColumnWidth(9, 60)
        self.tblCompilado.setColumnWidth(10, 80)
        self.tblCompilado.setColumnWidth(11, 90)

        self.load_data()

        # Conexión del QComboBox
        self.cbAlamcen.currentIndexChanged.connect(self.update_selected_value_label)

        #TEXTBOX BUSCAR
        self.txBuscarMateriales.textChanged.connect(self.filter_data)
        
        #BOTON PASAR MATERIALES
        self.btnPasarMateriales.clicked.connect(self.pasarDatosMateriales)

        #BOTON PASAR LETRAS
        self.btnPasarLetras.clicked.connect(self.pasarDatosLetras)


        #BOTON BUSCAR LETRAS
        self.btnBuscarLetras.clicked.connect(self.CargaCaracteres)

        #BOTON LIMPIAR
        self.btnLimpiar.clicked.connect(self.vaciar)

        #BOTON LIMPIAR TABLA LETRAS
        self.btnLimpiarLetras.clicked.connect(self.borrar)

        #BOTON COPIAR
        self.btnCopiar.clicked.connect(self.copiar)

        # Variable para guardar el índice de la última fila seleccionada
        self.selected_row_index = None

        # Configurar la señal para detectar cuando se selecciona una fila en la tabla 2
        self.tblCompilado.cellClicked.connect(self.on_row_selected)

        #BOTON ELIMINAR REGISTRO DE LA TABLA COMPILADO
        self.btnEliminar.clicked.connect(self.delete_selected_row)

        #BOTON EXPORTAR A EXCEL
        self.btnExcel.clicked.connect(self.ExportExcel)


        # Hacemos que todas las columnas de la tabla1, excepto la columna 1, se ajusten automáticamente
        for col in range(self.tblMateriales.columnCount()):
            if col != 0:
                self.tblMateriales.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
            else:
                self.tblMateriales.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.Fixed)
                self.tblMateriales.horizontalHeader().setSectionResizeMode(col, QHeaderView.ResizeMode.Interactive)
                self.tblMateriales.setColumnWidth(0, 70)
                self.tblMateriales.setColumnWidth(1, 300)
                

    def update_selected_value_label(self, index):
        self.almacen = self.cbAlamcen.currentText()

#FUNCION CARGA LOS MATERIALES DESDE UN ARCHIVO EXCEL
    def load_data(self):
        path = "materiales.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        
        self.tblMateriales.setRowCount(sheet.max_row)
        self.tblMateriales.setColumnCount(sheet.max_column)
        
        list_values = list(sheet.values)
        self.tblMateriales.setHorizontalHeaderLabels(list_values[0])
        
        row_index = 0
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                self.tblMateriales.setItem(row_index , col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1

#FUNCION PASA DATOS DE MATERIALES A LA TABLA COMPILADO
    def pasarDatosMateriales(self):
        # Obtener las filas seleccionadas de la tabla de origen
        selected_rows = set(index.row() for index in self.tblMateriales.selectedIndexes())

        # Obtener el número de filas del QTableWidget de destino
        dest_rows = self.tblCompilado.rowCount()

        # Agregar los registros seleccionados al QTableWidget de destino debajo del último registro existente
        for row in selected_rows:
            value_col1 = self.tblMateriales.item(row, 0).text()
            value_col2 = self.tblMateriales.item(row, 1).text()
            value_col3 = self.tblMateriales.item(row, 2).text()
            exists = False
            for dest_row in range(dest_rows):
                if (
                    self.tblCompilado.item(dest_row, 0).text() == value_col1
                ):
                    exists = True
                    break
            if not exists:
                insert_row = dest_rows
                dest_rows += 1
                self.tblCompilado.setRowCount(dest_rows)
                self.tblCompilado.setItem(insert_row, 0, QTableWidgetItem(value_col1))
                self.tblCompilado.setItem(insert_row, 1, QTableWidgetItem(value_col2))
                self.tblCompilado.setItem(insert_row, 2, QTableWidgetItem(value_col3))
                self.tblCompilado.setItem(insert_row, 3, QTableWidgetItem(''))
                self.tblCompilado.setItem(insert_row, 4, QTableWidgetItem('L'))
                self.tblCompilado.setItem(insert_row, 5, QTableWidgetItem(''))
                self.tblCompilado.setItem(insert_row, 6, QTableWidgetItem('P001'))
                self.tblCompilado.setItem(insert_row, 7, QTableWidgetItem(self.almacen))
                self.tblCompilado.setItem(insert_row, 8, QTableWidgetItem('0010'))
                self.tblCompilado.setItem(insert_row, 9, QTableWidgetItem('VALORADO'))
                self.tblCompilado.setItem(insert_row, 10, QTableWidgetItem(''))
                self.tblCompilado.setItem(insert_row, 11, QTableWidgetItem(''))
                value_col3 = ''
                self.load_data()
                return self.statusbar.showMessage('')
                
            else:
                return self.statusbar.showMessage('Este material ya fue agregado')
        value_col3 = '' 
        

 
#FUNCION PASA DATOS LETRAS A LA TABLA COMPILADO
    def pasarDatosLetras(self):
        # Obtener el número de filas del QTableWidget de origen
        num_filas_origen = self.tblLetras.rowCount()

        # Obtener el número de filas del QTableWidget de destino
        num_filas_destino = self.tblCompilado.rowCount()

        # Recorrer todas las filas de la tabla de origen
        for row in range(num_filas_origen):
            value_col1 = self.tblLetras.item(row, 0).text()
            value_col2 = self.tblLetras.item(row, 1).text()
            value_col3 = self.tblLetras.item(row, 2).text()

            # Verificar si el registro ya existe en la tabla de destino
            exists = False
            for dest_row in range(num_filas_destino):
                if (
                    self.tblCompilado.item(dest_row, 0).text() == value_col1
                ):
                    exists = True
                    break

            # Agregar el registro a la tabla de destino si no existe
            if not exists:
                self.tblCompilado.insertRow(num_filas_destino)
                self.tblCompilado.setItem(num_filas_destino, 0, QTableWidgetItem(value_col1))
                self.tblCompilado.setItem(num_filas_destino, 1, QTableWidgetItem(value_col2))
                self.tblCompilado.setItem(num_filas_destino, 2, QTableWidgetItem(value_col3))
                self.tblCompilado.setItem(num_filas_destino, 3, QTableWidgetItem(''))
                self.tblCompilado.setItem(num_filas_destino, 4, QTableWidgetItem('L'))
                self.tblCompilado.setItem(num_filas_destino, 5, QTableWidgetItem(''))
                self.tblCompilado.setItem(num_filas_destino, 6, QTableWidgetItem('P001'))
                self.tblCompilado.setItem(num_filas_destino, 7, QTableWidgetItem(self.almacen))
                self.tblCompilado.setItem(num_filas_destino, 8, QTableWidgetItem('0010'))
                self.tblCompilado.setItem(num_filas_destino, 9, QTableWidgetItem('VALORADO'))
                self.tblCompilado.setItem(num_filas_destino, 10, QTableWidgetItem(''))
                self.tblCompilado.setItem(num_filas_destino, 11, QTableWidgetItem(''))
                #return self.statusbar.showMessage('')
            else:
                return self.statusbar.showMessage('Favor volver a compilar.')           


                
    def filter_data(self):
        search_text = self.txBuscarMateriales.text().lower()
        for row in range(self.tblMateriales.rowCount()):
            item_col1 = self.tblMateriales.item(row, 0)
            item_col2 = self.tblMateriales.item(row, 1)
            if item_col1 is not None and item_col2 is not None:
                text_col1 = item_col1.text().lower()
                text_col2 = item_col2.text().lower()
                if search_text in text_col1 or search_text in text_col2:
                    self.tblMateriales.setRowHidden(row, False)
                else:
                    self.tblMateriales.setRowHidden(row, True)

#FUNCION BORRAR 
    def borrar(self):
        c_fila = self.tblLetras.rowCount() 
        while c_fila > 0:
            c_fila-= 1
            self.tblLetras.removeRow(c_fila)

#FUNCION LIMPIAR
    def vaciar(self):
        c_fila = self.tblCompilado.rowCount() 
        while c_fila > 0:
            c_fila-= 1
            self.tblCompilado.removeRow(c_fila)

        self.borrar()
        self.load_data()
        self.txBuscarLetras.clear()
        self.txBuscarMateriales.clear()
        self.statusbar.showMessage('')


#FUNCION BORRA REGISTROS DE LA TABLA COMPILADO
    def delete_selected_row(self):
        if self.selected_row_index is not None:
            self.tblCompilado.removeRow(self.selected_row_index)
            self.selected_row_index = None

    def on_row_selected(self, row_index):
        self.selected_row_index = row_index
        

#FUNCION CARGAR LETRAS
    def CargaCaracteres(self):
        input_text = self.txBuscarLetras.text()
        if not input_text:
            self.statusbar.showMessage('Favor escribir las facilidades!')
            return

        letras = {'A': '1000543', 'B': '1000542', 'C': '1000532', 'D': '1000525', 'E': '1000526', 'F': '1000527', 'G': '1000528', 'H': '1000529',
                  'I': '1000530', 'J': '1000531', 'K': '1000533', 'L': '1000541', 'M': '1000534', 'N': '1000535', 'O': '1000536', 'P': '1001185',
                  'Q': '1001186', 'R': '1001187', 'S': '1001188', 'T': '1001189', 'U': '1001190', 'V': '1001191', 'W': '1001192', 'X': '1001193',
                  'Y': '1001194', 'Z': '1001195', '1': '1001196', '2': '1001197', '3': '1000970', '4': '1001262', '5': '1000968', '6': '1000784',
                  '7': '1000670', '8': '1000682', '9': '1000784', '0': '1000536', '/': 'N/A', '*': 'N/A', '-': 'N/A', '+': 'N/A', '.': 'N/A',
                  '_': 'N/A', '(': 'N/A', ')': 'N/A', '=': 'N/A', '^': 'N/A', '&': 'N/A', '%': 'N/A', '$': 'N/A', '#': 'N/A', '@': 'N/A',
                  '~': 'N/A', '|': 'N/A', '<': 'N/A', '>': 'N/A', ',': 'N/A', '?': 'N/A', ';': 'N/A', ':': 'N/A', ' ': 'N/A', '\\': 'N/A',
                  '{': 'N/A', '}': 'N/A', '?': 'N/A', '[': 'N/A', ']': 'N/A', '"': 'N/A'}

        self.borrar()
        self.statusbar.showMessage('')
        self.noAplican = []
        self.datos = []

        # Count the occurrences of characters in the input text
        conteo = {char: input_text.count(char) for char in set(input_text)}

        for char, count in conteo.items():
            le = letras.get(char.upper(), 'N/A')
            if le == 'N/A':
                self.noAplican.append((letras[char.upper()], char.upper(), str(count)))
            else:
                self.datos.append((letras[char.upper()], char.upper(), str(count)))

        # Now, update the table widget
        for row, registro in enumerate(self.datos):
            self.tblLetras.insertRow(row)
            for column, elemento in enumerate(registro):
                cell = QTableWidgetItem(elemento)
                self.tblLetras.setItem(row, column, cell)

            

#FUNCION COPIAR DATOS DE LA TABLA COMPILADO
    def copiar(self):
        # Obtener los datos de la tabla 2
        data = []
        for row in range(self.tblCompilado.rowCount()):
            row_data = []
            for col in range(self.tblCompilado.columnCount()):
                item = self.tblCompilado.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append("")
            data.append(row_data)

        # Convertir los datos en una cadena de texto con formato de tabla
        text = ""
        for row in data:
            text += "\t".join(row) + "\n"

        # Copiar la cadena de texto al portapapeles
        try:
            cp.copy(text)
            # Mostrar mensaje de éxito
            return self.statusbar.showMessage('Los datos se han copiado al portapapeles.')
        except cp.PyperclipException:
            return self.statusbar.showMessage(' No se puede copiar al portapapeles.')
     

        
#EXPORTAR A EXCEL
    def ExportExcel(self):
        rowCount = self.tblCompilado.rowCount()
        if rowCount == 0:
            return self.statusbar.showMessage('La tabla no contiene datos!')
        columnCount = 12

        # add this line
        data = []

        for row in range(rowCount):
            rowData = []
            for column in range(columnCount):
                widgetItem = self.tblCompilado.item(row, column)
                if widgetItem and widgetItem.text:
                    rowData.append(widgetItem.text())

            # add this line
            data.append(rowData)
        # change these two lines
        df = pd.DataFrame(data)
        df.to_excel('MaterialesCompilado.xlsx', header=False, index=False)
        return self.statusbar.showMessage('Datos exportados con exito!')
    



if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec())
